import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import sys
import re

DEFAULT_EXCEL_FILE = r'D:\Projects\PaySlipGen\payroll_data.xlsx'
DEFAULT_OUTPUT_PDF = r'D:\Projects\PaySlipGen\payslips.pdf'

# ── Column header keywords ──────────────────────────────────────────────────
# Each entry: internal_key -> list of keywords that must ALL appear in the header
# (case-insensitive). Adjust keywords here if column names change in future sheets.
COLUMN_MAP = {
    'no':                    ['no'],
    'emp_no':                ['employe', 'no'],
    'name':                  ['employe', 'name'],
    'location':              ['location'],
    'total_days':            ['total', 'days'],
    'basic_salary':          ['basic', 'salary'],
    'variable_allowance':    ['variable', 'attendance', 'allowance'],
    'ot_hours':              ['over', 'time'],
    'hard_works_allowance':  ['hard', 'works'],
    'gang_leader_allowance': ['gang', 'leader'],
    'gross_pay':             ['gross', 'pay'],
    'epf_employee':          ['epf', 'employee'],
    'stamp_duty':            ['stamp'],
    'net_salary':            ['net', 'salary'],
    'epf_employer':          ['epf', 'employer', '12'],
    'etf_employer':          ['etf'],
    'total_epf':             ['total', 'epf'],
    'bank':                  ['name', 'bank'],
    'branch':                ['branch'],
    'account_no':            ['account'],
    'nic':                   ['nic'],
    'tel':                   ['tel'],
    'transaction_id':        ['transaction'],
    'payment_date':          ['payment', 'date'],
}


def get_payroll_sheet_names(excel_file):
    """
    Return a list of sheet names that contain payroll employee data.

    A sheet qualifies only if find_header_row_and_map succeeds on it —
    i.e. it has all essential columns (Employe Name, Gross Pay, Net Salary,
    Basic Salary).  This correctly excludes BC/Invoice/summary sheets even
    though those sheets contain the word "Employee" inside EPF column headers.
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    payroll_sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            find_header_row_and_map(ws)   # raises ValueError if not a payroll sheet
            payroll_sheets.append(sheet_name)
        except ValueError:
            pass
    return payroll_sheets


def find_header_row_and_map(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        non_none = [v for v in row if v is not None]
        if len(non_none) < 6:
            continue
        col_index = {}
        for key, keywords in COLUMN_MAP.items():
            for col_i, cell_val in enumerate(row):
                if cell_val is None:
                    continue
                cell_str = str(cell_val).lower()
                if all(kw in cell_str for kw in keywords):
                    if key not in col_index:
                        col_index[key] = col_i
        essential = {'name', 'gross_pay', 'net_salary', 'basic_salary'}
        if essential.issubset(col_index.keys()):
            print(f"  Header found at row {row_idx}")
            return row_idx, col_index
    raise ValueError(
        "Could not find a valid header row. "
        "Check that the sheet has columns like 'Employe Name', 'Gross Pay', 'Net Salary', etc."
    )


def parse_int_like(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    try:
        return int(text)
    except (ValueError, TypeError):
        try:
            as_float = float(text)
            return int(as_float) if as_float.is_integer() else None
        except (ValueError, TypeError):
            return None


def parse_float_like(value, default=None):
    if value is None or isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return default

    try:
        return float(text)
    except (ValueError, TypeError):
        return default

def is_blank(value):
    return value is None or str(value).strip() == ""

def build_validation_report(summary):
    lines = [
        "# Payslip Validation Report",
        "",
        f"- Source file : {summary['excel_file']}",
        f"- Sheet       : {summary['sheet_name']}",
        f"- Output PDF  : {summary['output_pdf']}",
        f"- Date range  : {summary['date_range']}",
        f"- Rows scanned: {summary['rows_scanned']}",
        f"- Candidates  : {summary['candidate_rows']}",
        f"- Generated   : {summary['employees_processed']}",
        f"- Issues      : {len(summary['issues'])}",
        "",
    ]

    if summary['issues']:
        lines += [
            "## Issues", "",
            "| Row | Field | Reason | Raw Value |",
            "|---|---|---|---|",
        ]
        for issue in summary['issues']:
            raw_text = str(issue['raw_value']).replace("|", "\\|")
            lines.append(
                f"| {issue['row']} | {issue['field']} | {issue['reason']} | {raw_text} |"
            )
    else:
        lines.append("No validation issues were detected.")

    return "\n".join(lines) + "\n"

def extract_employees(ws, header_row_idx, col_index):
    employees = []
    issues = []
    rows_scanned = 0
    candidate_rows = 0

    numeric_fields = [
        'total_days', 'basic_salary', 'variable_allowance', 'ot_hours',
        'hard_works_allowance', 'gang_leader_allowance', 'gross_pay',
        'epf_employee', 'stamp_duty', 'net_salary', 'epf_employer',
        'etf_employer', 'total_epf',
    ]

    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 2, max_row=ws.max_row, values_only=True),
        start=header_row_idx + 2,
    ):
        rows_scanned += 1

        def get(key):
            idx = col_index.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        no_raw = get('no')
        if not is_blank(no_raw):
            candidate_rows += 1

        no_int = parse_int_like(no_raw)
        if no_int is None:
            if not is_blank(no_raw):
                issues.append({
                    'row': row_number, 'field': 'no',
                    'reason': 'Invalid employee number (must be an integer)',
                    'raw_value': no_raw,
                })
            continue

        name = str(get('name') or '').strip()
        if not name:
            issues.append({
                'row': row_number, 'field': 'name',
                'reason': 'Missing employee name', 'raw_value': get('name'),
            })

        for field in numeric_fields:
            raw_val = get(field)
            if is_blank(raw_val):
                continue
            if parse_float_like(raw_val, default=None) is None:
                issues.append({
                    'row': row_number, 'field': field,
                    'reason': 'Expected numeric value', 'raw_value': raw_val,
                })

        employees.append({
            'no': no_int,
            'emp_no': get('emp_no'),
            'name': name,
            'location': str(get('location') or '').strip(),
            'total_days': get('total_days'),
            'basic_salary': get('basic_salary'),
            'variable_allowance': get('variable_allowance'),
            'ot_hours': get('ot_hours'),
            'hard_works_allowance': get('hard_works_allowance'),
            'gang_leader_allowance': get('gang_leader_allowance'),
            'gross_pay': get('gross_pay'),
            'epf_employee': get('epf_employee'),
            'stamp_duty': get('stamp_duty'),
            'net_salary': get('net_salary'),
            'epf_employer': get('epf_employer'),
            'etf_employer': get('etf_employer'),
            'total_epf': get('total_epf'),
            'bank': str(get('bank') or '').strip(),
            'branch': str(get('branch') or '').strip(),
            'account_no': get('account_no'),
            'nic': get('nic'),
            'tel': get('tel'),
            'transaction_id': get('transaction_id'),
            'payment_date': get('payment_date'),
        })
    return employees, issues, rows_scanned, candidate_rows


def extract_date_range(ws, max_rows=30):
    """Scan top rows for a date range like 2026.02.28 - 2026.03.06."""
    range_pattern = re.compile(
        r"(\d{4}[./-]\d{1,2}[./-]\d{1,2})\s*-\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2})"
    )
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for val in row:
            if val is None:
                continue
            cell_str = str(val).strip()
            if not cell_str:
                continue
            match = range_pattern.search(cell_str)
            if match:
                return f"{match.group(1)} - {match.group(2)}"
    return "N/A"


def _resolve_sheet(wb, sheet_name):
    """
    Return the worksheet to use.
    If sheet_name is given and exists, use it.
    Otherwise scan all sheets for the first one with an employee header.
    Falls back to the active sheet if nothing is found.
    """
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]

    for sname in wb.sheetnames:
        candidate = wb[sname]
        for row in candidate.iter_rows(min_row=1, max_row=20, values_only=True):
            non_none = [v for v in row if v is not None]
            if len(non_none) >= 6 and any(
                'employe' in str(v).lower() for v in non_none if v
            ):
                return candidate

    return wb.active  # last resort


def fmt(val):
    as_num = parse_float_like(val, default=None)
    if as_num is None:
        return "0.00" if val is None else str(val)
    return f"{as_num:,.2f}"

def draw_payslip(c, emp, date_range, reg_no="NOT SET"):
    W, H = A4
    margin = 15 * mm
    w = W - 2 * margin
    y = H - margin

    c.setFillColor(colors.HexColor("#1a3a5c"))
    c.rect(margin, y - 22*mm, w, 22*mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(W/2, y - 10*mm, "ReliaPro Manpower")
    c.setFont("Helvetica", 9)
    c.drawCentredString(W/2, y - 16*mm, "Siyambalanduwa")
    y -= 24*mm

    c.setFillColor(colors.HexColor("#2e6da4"))
    c.rect(margin, y - 9*mm, w, 9*mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(W/2, y - 6*mm, f"PAY SLIP  |  Period: {date_range}")
    y -= 11*mm

    c.setStrokeColor(colors.HexColor("#2e6da4"))
    c.setLineWidth(0.5)
    c.setFillColor(colors.HexColor("#f0f5fb"))
    c.rect(margin, y - 28*mm, w, 28*mm, fill=1, stroke=1)

    lx = margin + 4*mm
    rx = margin + w/2 + 2*mm

    def info_row(label, value, x, row_y):
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(colors.HexColor("#555555"))
        c.drawString(x, row_y, label)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)
        c.drawString(x + 28*mm, row_y, str(value) if value else "-")

    info_row("Employee Name", emp['name'],     lx, y - 7*mm)
    info_row("Employee No.",  emp['emp_no'],   lx, y - 13*mm)
    info_row("NIC Number",    emp['nic'],       lx, y - 19*mm)
    info_row("Location",      emp['location'] or "Siyambalanduwa", lx, y - 25*mm)
    info_row("Bank",          emp['bank'],      rx, y - 7*mm)
    info_row("Branch",        emp['branch'],    rx, y - 13*mm)
    info_row("Account No.",   emp['account_no'],rx, y - 19*mm)
    info_row("Payment Date",  emp['payment_date'], rx, y - 25*mm)
    y -= 30*mm

    # ── Table helpers ────────────────────────────────────────────────────────
    def section_header(label, sy):
        c.setFillColor(colors.HexColor("#2e6da4"))
        c.rect(margin, sy - 7*mm, w, 7*mm, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(margin + 3*mm, sy - 4.5*mm, label)
        return sy - 7*mm

    def table_row(label, value, sy, shade=False):
        row_h = 6.5*mm
        if shade:
            c.setFillColor(colors.HexColor("#f7f9fc"))
            c.rect(margin, sy - row_h, w, row_h, fill=1, stroke=0)
        c.setStrokeColor(colors.HexColor("#dddddd"))
        c.setLineWidth(0.3)
        c.line(margin, sy - row_h, margin + w, sy - row_h)
        c.setFont("Helvetica", 8.5)
        c.setFillColor(colors.HexColor("#333333"))
        c.drawString(margin + 3*mm, sy - 4.5*mm, label)
        c.setFont("Helvetica-Bold", 8.5)
        c.setFillColor(colors.black)
        c.drawRightString(margin + w - 3*mm, sy - 4.5*mm, f"LKR {value}")
        return sy - row_h

    # ── Earnings ─────────────────────────────────────────────────────────────
    y = section_header("EARNINGS", y)
    y = table_row("Basic Salary", fmt(emp['basic_salary']), y, True)
    y = table_row("Variable Attendance Allowance", fmt(emp['variable_allowance']), y, False)
    y = table_row(
        f"Variable Hard Works Allowance  (OT: {emp['ot_hours'] or 0} hrs)",
        fmt(emp['hard_works_allowance']), y, True,
    )
    if parse_float_like(emp['gang_leader_allowance'], default=0.0) > 0:
        y = table_row("Variable Allowance (Gang Leader)", fmt(emp['gang_leader_allowance']), y, False)

    # Gross Pay highlight
    c.setFillColor(colors.HexColor("#e8f0fb"))
    c.rect(margin, y - 7*mm, w, 7*mm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#2e6da4"))
    c.setLineWidth(0.8)
    c.rect(margin, y - 7*mm, w, 7*mm, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(colors.HexColor("#1a3a5c"))
    c.drawString(margin + 3*mm, y - 4.5*mm, "GROSS PAY")
    c.drawRightString(margin + w - 3*mm, y - 4.5*mm, f"LKR {fmt(emp['gross_pay'])}")
    y -= 9*mm

    # ── Deductions ────────────────────────────────────────────────────────────
    y = section_header("DEDUCTIONS", y)
    y = table_row("EPF Employee Contribution (8%)", fmt(emp['epf_employee']), y, True)
    y = table_row("Stamp Duty", fmt(emp['stamp_duty']), y, False)

    # Net Salary highlight
    c.setFillColor(colors.HexColor("#1a3a5c"))
    c.rect(margin, y - 8*mm, w, 8*mm, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.white)
    c.drawString(margin + 3*mm, y - 5.5*mm, "NET SALARY")
    c.drawRightString(margin + w - 3*mm, y - 5.5*mm, f"LKR {fmt(emp['net_salary'])}")
    y -= 10*mm

    # ── EPF / ETF ─────────────────────────────────────────────────────────────
    y = section_header("EPF / ETF INFORMATION", y)
    y = table_row("EPF Employer Contribution (12%)", fmt(emp['epf_employer']), y, True)
    y = table_row("ETF Employer Contribution (3%)",  fmt(emp['etf_employer']), y, False)
    y = table_row("Total EPF Contribution (20%)",    fmt(emp['total_epf']),    y, True)
    y -= 3*mm

    if emp['transaction_id']:
        c.setFont("Helvetica", 7.5)
        c.setFillColor(colors.HexColor("#777777"))
        c.drawString(margin, y - 4*mm, f"Transaction ID: {emp['transaction_id']}")
        y -= 6*mm

    # ── Signatures ────────────────────────────────────────────────────────────
    y -= 5*mm
    sig_y = y - 8*mm
    c.setStrokeColor(colors.HexColor("#999999"))
    c.setLineWidth(0.5)
    c.line(margin, sig_y, margin + 55*mm, sig_y)
    c.line(margin + w - 55*mm, sig_y, margin + w, sig_y)
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#555555"))
    c.drawCentredString(margin + 27*mm,     sig_y - 4*mm, "Employee Signature")
    c.drawCentredString(margin + w - 27*mm, sig_y - 4*mm, "Managing Director Signature")

    # ── Footer ────────────────────────────────────────────────────────────────
    c.setFillColor(colors.HexColor("#1a3a5c"))
    c.rect(margin, margin, w, 6*mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 7)
    c.drawCentredString(W/2, margin + 2*mm,
                        f"Reg No: {reg_no} | This is a computer generated pay slip")


def generate_payslips(excel_file, output_pdf, write_validation_report=True,
                      reg_no="NOT SET", sheet_name=None):
    """
    Generate payslips from a single sheet.

    Parameters
    ----------
    excel_file : str
        Path to the .xlsx workbook.
    output_pdf : str
        Destination PDF path.
    write_validation_report : bool
        If True, writes a .md report next to the PDF.
    reg_no : str
        Company registration number shown in the payslip footer.
    sheet_name : str or None
        Name of the sheet to process.  If None, the first payroll sheet found
        is used (same behaviour as before this parameter was added).
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = _resolve_sheet(wb, sheet_name)
    resolved_sheet = ws.title
    print(f"Processing sheet: '{resolved_sheet}'")

    date_range = extract_date_range(ws)
    print(f"  Date range : {date_range}")

    header_row_idx, col_index = find_header_row_and_map(ws)
    employees, issues, rows_scanned, candidate_rows = extract_employees(
        ws, header_row_idx, col_index
    )
    print(f"  Employees  : {len(employees)}")

    c = canvas.Canvas(output_pdf, pagesize=A4)
    for emp in employees:
        draw_payslip(c, emp, date_range, reg_no)
        c.showPage()
    c.save()

    summary = {
        'excel_file':          excel_file,
        'sheet_name':          resolved_sheet,
        'output_pdf':          output_pdf,
        'date_range':          date_range,
        'rows_scanned':        rows_scanned,
        'candidate_rows':      candidate_rows,
        'employees_processed': len(employees),
        'issues':              issues,
        'validation_report_path': None,
    }

    if write_validation_report:
        report_path = f"{output_pdf}.validation_report.md"
        with open(report_path, "w", encoding="utf-8") as fp:
            fp.write(build_validation_report(summary))
        summary['validation_report_path'] = report_path
        print(f"  Report     : {report_path}")

    print(f"  Saved      : {output_pdf}")
    return summary


def main():
    excel_file = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_FILE
    output_pdf = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_PDF
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
    generate_payslips(excel_file, output_pdf, sheet_name=sheet_name)


if __name__ == "__main__":
    main()